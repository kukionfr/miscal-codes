import glob
import os
import sys
from PIL import Image, ImageDraw
from openpyxl import load_workbook

# collect all the coordinates from the first
coord_list = []

# x0, y0, x1, y1 in order
coord = [0, 0, 0, 0]

# data starts from second row
startrow = 2 

# the diameter of circle to be drawn
diamater = 3

# the number of frames in the current movie
numofframes = len(glob.glob("./Full Movie/*.tif"))
print "The total number of frames in this movie is: ", numofframes, "\n"

# display where this file exists
print "Starting Path is : ", os.getcwd(), "\n"

# change path to deal with excel files 
os.chdir("./ExcelFiles")
print "Path changed to where xlsx files are at: ", os.getcwd(), "\n"

parasites = glob.glob('./*Track_[0-9].xlsx')
print 'Total number of parasite under tracking is: ', len(parasites), "\n"

#preallocate the S!!
paradata = range (len(parasites))

# paratites' no.
firstindex = 0

# info : secondindex : parasite n's frame np.

for parafile in parasites :
	wb = load_workbook(filename = parafile)
	ws = wb['Sheet1']
	lenofrow = ws.max_row-1
	paradata[firstindex] = range(lenofrow)
	# insert frame data with ws
	for secondindex in range(len(paradata[firstindex])):
		frame =  int(ws.cell(row = secondindex+2, column = 1).value)
		x =  int(ws.cell(row = secondindex+2, column = 7).value)
		y =  int(ws.cell(row = secondindex+2, column = 8).value)
		paradata[firstindex][secondindex] = [frame, x, y]
	firstindex = firstindex + 1

for i in paradata:
	print i,"length is",len(i),"\n"

os.chdir("../Full Movie")

frameno = 0
colorset = ["red", "yellow", "green", "white", "blue", "cyan", "pink", "sienna"]

for frame in glob.glob("*.tif"):
	im = Image.open(frame)
	rgbimg = Image.new("RGBA", im.size)
	rgbimg.paste(im)
	for firstindex in range(len(paradata)):
		for secondindex in range(len(paradata[firstindex])):
			if(paradata[firstindex][secondindex][0] == (frameno+1)):
					print "====== Process", frameno + 1, "======"
					# Create and assign Draw instance
					draw = ImageDraw.Draw(rgbimg)
					# compare the frame number with other parasite's last frame
					for paraid in range(len(paradata)):
						#is last frame number is less than this frame?
						if paradata[paraid][-1][0] < frameno+1 :
							linecolor = colorset[paraid]
							#then draw the all the previous points of each parasite
							#print "Yes my life is over... ", paraid
							for dataid in range(len(paradata[paraid])):
								#print "I am drawing other parasite frame #", dataid
								# set the coordination of the box to be drawn
								coord[0] = paradata[paraid][dataid][1] - diamater/2
								coord[1] = paradata[paraid][dataid][2] - diamater/2
								coord[2] = paradata[paraid][dataid][1] + diamater/2
						 		coord[3] = paradata[paraid][dataid][2] + diamater/2
				 				# a circle is drawn inside the box
								# outline parameter : color. linecolor
								draw.ellipse(coord, outline=linecolor)

					linecolor = colorset[firstindex]

					if(secondindex == 0):
						# set the coordination of the box to be drawn
						coord[0] = paradata[firstindex][secondindex][1] - diamater/2
						coord[1] = paradata[firstindex][secondindex][2] - diamater/2
						coord[2] = paradata[firstindex][secondindex][1] + diamater/2
				 		coord[3] = paradata[firstindex][secondindex][2] + diamater/2
				 		# a circle is drawn inside the box
						# outline parameter : color. linecolor
						draw.ellipse(coord, outline=linecolor)
					for previndex in range(secondindex):
						# set the coordination of the box to be drawn
						coord[0] = paradata[firstindex][previndex][1] - diamater/2
						coord[1] = paradata[firstindex][previndex][2] - diamater/2
						coord[2] = paradata[firstindex][previndex][1] + diamater/2
				 		coord[3] = paradata[firstindex][previndex][2] + diamater/2
						# a circle is drawn inside the box
						# outline parameter : color. linecolor
						draw.ellipse(coord, outline=linecolor)
					# save with the new name in the same extension
					file, ext = os.path.splitext(frame)
					rgbimg.save(file + "_modified.tif", "TIFF")
					# open modified file and check
					Image.open(file + "_modified.tif")
					#rgbimg.show()
	frameno = frameno + 1
